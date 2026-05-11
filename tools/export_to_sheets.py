"""
Export the top-50 lead list and RM briefings to a formatted Excel workbook.

Output: .tmp/leads_report_YYYYMMDD.xlsx
  Leads        — 50 leads with score tiers colour-coded (A=green, B=yellow, C=orange)
  RM_Briefings — talking points per lead (merged from rm_briefings.csv if available)
  Audit_Log    — one row per export, append-only within the same file

Run this after select_top_leads.py (and optionally generate_rm_briefing.py).
The file can be emailed to RMs or placed on a shared drive.
"""

import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
TMP_DIR = PROJECT_ROOT / ".tmp"
INPUT_PATH = TMP_DIR / "top50_leads.csv"
BRIEFINGS_PATH = TMP_DIR / "rm_briefings.csv"

TIER_FILLS = {
    "A": PatternFill("solid", fgColor="C6EFCE"),   # green
    "B": PatternFill("solid", fgColor="FFEB9C"),   # yellow
    "C": PatternFill("solid", fgColor="FFCCBA"),   # orange
}
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

LEADS_COLS = {
    "lead_id":    "Lead_ID",
    "lead_score": "Score",
    "score_tier": "Score_Tier",
    "age":        "Age",
    "job":        "Job",
    "education":  "Education",
    "marital":    "Marital",
    "balance":    "Balance_EUR",
    "housing":    "Housing_Loan",
    "loan":       "Personal_Loan",
    "duration":   "Engagement_s",
    "campaign":   "Campaign_Contacts",
    "poutcome":   "Prev_Outcome",
    "rm_assigned": "RM_Assigned",
    "status":     "Status",
    "notes":      "Notes",
    "last_updated": "Last_Updated",
}

BRIEFINGS_COLS = {
    "lead_id":      "Lead_ID",
    "rm_assigned":  "RM_Assigned",
    "score_tier":   "Score_Tier",
    "why_prospect": "Why_Prospect",
    "opening_line": "Opening_Line",
    "talking_point": "Talking_Point",
    "briefing_parse_error": "Parse_Error",
}


def _write_header(ws, headers: list[str]):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


def _write_rows(ws, rows: list[list], start_row: int = 2, tier_col_idx: int | None = None):
    for r_idx, row in enumerate(rows, start=start_row):
        tier = row[tier_col_idx - 1] if tier_col_idx else None
        fill = TIER_FILLS.get(tier)
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill


def _autofit(ws, min_width: int = 10, max_width: int = 50):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


def write_leads_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Leads")
    headers = list(LEADS_COLS.values())
    _write_header(ws, headers)

    tier_col_idx = headers.index("Score_Tier") + 1

    rows = []
    for _, row in df.iterrows():
        r = []
        for csv_col in LEADS_COLS:
            val = row.get(csv_col, "")
            if csv_col == "balance" and val != "":
                val = round(float(val), 2)
            elif csv_col == "lead_score" and val != "":
                val = round(float(val), 1)
            r.append(val if pd.notna(val) else "")
        rows.append(r)

    _write_rows(ws, rows, tier_col_idx=tier_col_idx)
    _autofit(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    print(f"  Sheet 'Leads': {len(rows)} rows written")


def write_briefings_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("RM_Briefings")
    headers = list(BRIEFINGS_COLS.values())
    _write_header(ws, headers)

    tier_col_idx = headers.index("Score_Tier") + 1

    rows = []
    for _, row in df.iterrows():
        r = [row.get(csv_col, "") for csv_col in BRIEFINGS_COLS]
        r = ["" if pd.isna(v) else v for v in r]
        rows.append(r)

    _write_rows(ws, rows, tier_col_idx=tier_col_idx)
    _autofit(ws, max_width=60)
    print(f"  Sheet 'RM_Briefings': {len(rows)} rows written")


def write_audit_sheet(wb: Workbook, action: str, detail: str, output_path: Path):
    ws = wb.create_sheet("Audit_Log")
    headers = ["Timestamp", "Action", "Detail"]
    _write_header(ws, headers)

    # If file already existed, load prior audit rows
    prior_rows = []
    if output_path.exists():
        try:
            old_wb = load_workbook(output_path)
            if "Audit_Log" in old_wb.sheetnames:
                old_ws = old_wb["Audit_Log"]
                for row in old_ws.iter_rows(min_row=2, values_only=True):
                    if any(row):
                        prior_rows.append(list(row))
        except Exception:
            pass

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    all_rows = prior_rows + [[timestamp, action, detail]]
    _write_rows(ws, all_rows)
    _autofit(ws)
    print(f"  Sheet 'Audit_Log': {len(all_rows)} entries")


def main():
    if not INPUT_PATH.exists():
        print(f"ERROR: {INPUT_PATH} not found. Run select_top_leads.py first.", file=sys.stderr)
        sys.exit(1)

    df = pd.read_csv(INPUT_PATH)
    print(f"Loaded {len(df)} leads from {INPUT_PATH.name}")

    now = datetime.now()
    week_num = now.strftime("%W")
    filename = f"leads_report_{now.strftime('%Y%m%d')}.xlsx"
    output_path = TMP_DIR / filename

    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    write_leads_sheet(wb, df)

    if BRIEFINGS_PATH.exists():
        briefings_df = pd.read_csv(BRIEFINGS_PATH)
        write_briefings_sheet(wb, briefings_df)
    else:
        print(f"  Skipping RM_Briefings: {BRIEFINGS_PATH.name} not found")

    write_audit_sheet(
        wb,
        action="LEADS_EXPORTED",
        detail=f"{len(df)} leads exported — Week {week_num}, {now.year}",
        output_path=output_path,
    )

    wb.save(output_path)
    print(f"\nSaved: {output_path}")
    print("Share this file with your RMs via email or shared drive.")


if __name__ == "__main__":
    main()
